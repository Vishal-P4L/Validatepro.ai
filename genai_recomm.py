import os
import streamlit as st
import pandas as pd
from frontend import *
from bs4 import BeautifulSoup
import spacy
import re
import openpyxl
import requests
from nltk.corpus import stopwords
from rake_nltk import Rake
import nltk
import PyPDF2
from docx import Document
import fitz  # PyMuPDF for PDF handling
from dotenv import load_dotenv
import google.generativeai as genai
import json



# Load environment variables from .env file
load_dotenv()

# Configure Generative AI API with the key
genai.configure(api_key=os.getenv("GOOGLE_API_KEY"))

# Initialize the Gemini model
model = genai.GenerativeModel("gemini-pro")
chat = model.start_chat(history=[])

# Ensure necessary directories are created
STATIC_DIRECTORY = "static"  

if not os.path.exists(STATIC_DIRECTORY):
    os.makedirs(STATIC_DIRECTORY)
    print(f"Directory '{STATIC_DIRECTORY}' created.")
else:
    print(f"Directory '{STATIC_DIRECTORY}' already exists.")

# Load spaCy's English language model
nlp = spacy.load("en_core_web_sm")

# Ensure NLTK stopwords are downloaded
nltk.download('stopwords', quiet=True)
nltk.download('averaged_perceptron_tagger', quiet=True)
nltk.download('punkt', quiet=True)
nltk_stopwords = set(stopwords.words('english'))

# Tokenization and keyword extraction functions
def normalize_text_spacy(text):
    doc = nlp(text)
    tokens = [token.lemma_ for token in doc if token.text.lower() not in nltk_stopwords and not token.is_punct]
    return tokens

def tokenize_column_as_list(column_values):
    tokenized_data = []
    for text in column_values:
        tokens = normalize_text_spacy(str(text))
        tokenized_data.append(tokens)
    return tokenized_data

def extract_keywords_rake(text):
    rake = Rake()
    rake.extract_keywords_from_text(text)
    ranked_phrases = rake.get_ranked_phrases()
    
    output_keywords = []
    for phrase in ranked_phrases:
        words = phrase.split()
        for i in range(len(words)):
            for j in range(i, len(words)):
                combination = ' '.join(words[i:j + 1])
                if len(combination.split()) >= 2:
                    output_keywords.append(combination)

    output_keywords = list(set(output_keywords))
    output_keywords.sort()
    
    return output_keywords

def match_keywords_in_config(keywords, config_data, config_file_name):
    matched_info = []
    for kw in keywords:
        for sheet_name, sheet_info in config_data.items():
            columns = sheet_info['columns']
            rows = sheet_info['rows']
            for row_idx, row_entries in enumerate(rows):
                for col_idx, entry in enumerate(row_entries):
                    if isinstance(entry, str) and kw.lower() in entry.lower():
                        column_name = columns[col_idx] if col_idx < len(columns) else f"Column {col_idx}"
                        matched_info.append(
                            {
                                "keyword": kw,
                                "sheet_name": sheet_name,
                                "row": row_idx+2,
                                "column": column_name,
                                "cell_value": entry  # Optional: store actual cell value for verification
                            }
                        )
    return matched_info

def read_pdf(file):
    pdf_reader = PyPDF2.PdfReader(file)
    text = []
    for page in pdf_reader.pages:
        page_text = page.extract_text()
        if page_text:  # Ensure we add only if there's text
            text.append(page_text)
    return ' '.join(text)  # Return as a single string

def read_docx(file):
    doc = Document(file)
    text = []
    for para in doc.paragraphs:
        text.append(para.text)
    return ' '.join(text)  # Return as a single string

# Function to safely split keywords
def safe_split_keywords(keywords):
    if isinstance(keywords, str) and keywords:
        return keywords.split('; ')
    return []




# Function to extract hyperlinks from a specific column
def extract_data_from_column(file_path, sheet_name, hyperlink_column):
    # Open the workbook and the specific sheet
    workbook = openpyxl.load_workbook(file_path, data_only=False)  # Set `data_only=False` to read formulas
    sheet = workbook[sheet_name]
    
    # Find the index of the specified column
    header_row = next(sheet.iter_rows(min_row=1, max_row=1))  # Assuming the first row is the header
    column_index = None
    for cell in header_row:
        if cell.value == hyperlink_column:
            column_index = cell.column
            break
    
    # Check if the column is found
    if column_index is None:
        raise ValueError(f"Column '{hyperlink_column}' not found in the sheet.")
    
    # Initialize a list to store hyperlink data
    hyperlink_data = []
    
    # Step 2: Iterate through the rows in the relevant column
    for row in sheet.iter_rows(min_row=2):  # Start from the second row to skip headers
        # Extract hyperlink from the specified column
        hyperlink_cell = row[column_index - 1]  # Adjust for zero-based indexing
        if hyperlink_cell.value and "HYPERLINK" in str(hyperlink_cell.value):  # Check if the cell contains a HYPERLINK formula
            match = re.search(r'HYPERLINK\("([^"]+)"', str(hyperlink_cell.value))
            if match:
                hyperlink_data.append(match.group(1))
        else:
            hyperlink_data.append(None)  # Placeholder if no hyperlink
    
    return hyperlink_data



# Streamlit UI
st.title("Data Filtering, Tokenization, and Keyword Matching Application")

# Upload main data file
uploaded_file = st.file_uploader("Upload a CSV or Excel file", type=['csv', 'xlsx'])

 
# Streamlit file upload and processing code
# Streamlit file upload and processing code
if uploaded_file:
    if uploaded_file.name.endswith('.csv'):
        df = pd.read_csv(uploaded_file)
    elif uploaded_file.name.endswith('.xlsx'):
        xlsx = pd.ExcelFile(uploaded_file)
        sheet_name = st.selectbox("Select a sheet", xlsx.sheet_names)
        df = pd.read_excel(uploaded_file, sheet_name=sheet_name)

        # Extract hyperlinks from the selected sheet
        try:
            hyperlink_column = "Release Note"  # Column containing hyperlinks
            extracted_data = extract_data_from_column(uploaded_file, sheet_name, hyperlink_column)

            # Convert the extracted data into a DataFrame with 'extracted_link' column
            extracted_df = pd.DataFrame(extracted_data, columns=["extracted_link"])

            # Add the extracted 'extracted_link' column to the original DataFrame (df)
            df["extracted_link"] = extracted_df["extracted_link"]

        except ValueError as e:
            st.write(e)

    st.write("Data Preview:")
    st.dataframe(df.head())

    # Filtering section
    columns = st.multiselect("Select columns for filtering", df.columns)
    if columns:
        selected_values = {}
        for column in columns:
            unique_values = df[column].dropna().unique()
            selected = st.multiselect(f"Select values to filter from '{column}'", unique_values.tolist())
            selected_values[column] = selected

        mask = pd.Series([True] * len(df))
        for column, values in selected_values.items():
            if values:
                mask &= df[column].isin(values)
        filtered_df = df[mask]
        st.write("Filtered Data with Selected Values:")
        st.write(f"Number of rows: {filtered_df.shape[0]}, Number of columns: {filtered_df.shape[1]}")
        st.dataframe(filtered_df)

        # Upload Configuration and URS files
        config_file = st.file_uploader("Upload Configuration File", type=["csv", "xlsx", "pdf", "docx"])
        urs_file = st.file_uploader("Upload URS File", type=["csv", "xlsx", "pdf", "docx"])

        def load_config_data(file):
            config_data = {}
            file_name = file.name

            if file_name.endswith('.xlsx'):
                config_xlsx = pd.ExcelFile(file)
                for sheet in config_xlsx.sheet_names:
                    sheet_data = config_xlsx.parse(sheet)
                    sheet_data.fillna('', inplace=True)
                    config_data[sheet] = {
                        'columns': sheet_data.columns.tolist(),
                        'rows': sheet_data.astype(str).values.tolist()
                    }
            elif file_name.endswith('.csv'):
                config_csv = pd.read_csv(file)
                config_csv.fillna('', inplace=True)
                config_data['CSV Sheet'] = {
                    'columns': config_csv.columns.tolist(),
                    'rows': config_csv.astype(str).values.tolist()
                }
            elif file_name.endswith('.pdf'):
                config_pdf_text = read_pdf(file)
                config_data['PDF Document'] = {
                    'columns': ['Content'],
                    'rows': [[config_pdf_text]]
                }
            elif file_name.endswith('.docx'):
                config_docx_text = read_docx(file)
                config_data['DOCX Document'] = {
                    'columns': ['Content'],
                    'rows': [[config_docx_text]]
                }
            return config_data

        if config_file:
            config_data = load_config_data(config_file)
        if urs_file:
            urs_data = load_config_data(urs_file)

        # Tokenization and Keyword Extraction for each file
        def process_both_files(filtered_df, config_data, urs_data):
            # Select column for tokenization only once
            column = st.selectbox("Select column for tokenization (Configuration and URS Files)", filtered_df.columns)
            column_values = filtered_df[column].dropna().values

            # Tokenization and keyword extraction for each text
            tokenized_lists = tokenize_column_as_list(column_values)
            extracted_keywords = [extract_keywords_rake(str(text)) for text in column_values]
            extracted_keywords_counts = [len(kw) for kw in extracted_keywords]

            # Add keywords columns to DataFrame
            filtered_df['Extracted Keywords'] = ['; '.join(kw) for kw in extracted_keywords]
            filtered_df['Extracted Keywords Count'] = pd.Series(extracted_keywords_counts, index=filtered_df.index)

            # Match keywords in configuration
            matched_keywords_details = filtered_df['Extracted Keywords'].apply(
                lambda x: match_keywords_in_config(safe_split_keywords(x), config_data, "Configuration File")
            )

            matched_keywords = []
            matched_details = []
            matched_sheet_names = []

            for details in matched_keywords_details:
                if details:
                    grouped_details = {}
                    for detail in details:
                        keyword = detail['keyword']
                        sheet_name = detail['sheet_name']
                        row = detail['row']
                        column = detail['column']

                        key = (keyword, sheet_name, column)
                        if key not in grouped_details:
                            grouped_details[key] = []
                        grouped_details[key].append(row)

                    formatted_details = []
                    unique_sheets = set()
                    for (keyword, sheet_name, column), rows in grouped_details.items():
                        formatted_rows = f"[{', '.join(map(str, rows))}]"
                        formatted_details.append(f"({keyword}: found in sheet name: {sheet_name}, rows: {formatted_rows}, column name: {column})")
                        unique_sheets.add(sheet_name)

                    matched_keywords.append('  ;  '.join({k[0] for k in grouped_details.keys()}))  # Unique keywords
                    matched_details.append('  ;   '.join(formatted_details))
                    matched_sheet_names.append('  ;  '.join(unique_sheets))  # Add unique sheets to the new column
                else:
                    matched_keywords.append("None") 
                    matched_details.append("None")
                    matched_sheet_names.append("None")

            # Add matched keywords, details, and sheet names to the DataFrame
            filtered_df['Matched Keywords'] = matched_keywords
            filtered_df['Matched Keywords Details'] = matched_details
            filtered_df['Unique Matched Sheet Names'] = matched_sheet_names

            # Display updated DataFrame
            st.write("Data with Matched Keywords and Details:")
            st.dataframe(filtered_df)

            # Option to download the updated DataFrame
            def convert_df_to_csv(df):
                return df.to_csv(index=False).encode('utf-8')

            csv_data = convert_df_to_csv(filtered_df)

            st.download_button(
                label="Download Filtered and Updated Data",
                data=csv_data,
                file_name='filtered_updated_data.csv',
                mime='text/csv',
            )

        # Process the files only if both are uploaded
        if config_file and urs_file:
            st.write("Processing both Configuration and URS files...")
            process_both_files(filtered_df, config_data, urs_data)
        elif config_file or urs_file:
            st.warning("Please upload both Configuration and URS files to proceed.")

        # Scrape data function
        def fetch_and_scrape(url):
                # Define a pattern for the desired URLs
                target_pattern = r"veevavault\.help"

                # Check if the URL matches the target pattern
                if re.search(target_pattern, url):
                    response = requests.get(url, verify=False)
                    if response.status_code == 200:
                        soup = BeautifulSoup(response.text, 'html.parser')
                        headings = soup.find_all(['h1', 'h2', 'h3', 'h4', 'h5', 'h6'])
                        result = []

                        for heading in headings:
                            content = {"paragraphs": [], "lists": []}
                            next_elem = heading.find_next_sibling()

                            while next_elem and next_elem.name not in ['h1', 'h2', 'h3', 'h4', 'h5', 'h6']:
                                if next_elem.name == 'p':
                                    content['paragraphs'].append(next_elem.get_text(strip=True))

                                if next_elem.name in ['ul', 'ol']:
                                    list_items = [li.get_text(strip=True) for li in next_elem.find_all('li')]
                                    content['lists'].extend(list_items)

                                next_elem = next_elem.find_next_sibling()

                            result.append({
                                "heading": heading.get_text(strip=True),
                                "content": content
                            })

                        return result
                    else:
                        print(f"Failed to fetch data. HTTP Status Code: {response.status_code}")
                        return None
                else:
                    print(f"URL does not match the target pattern: {url}")
                    return None

    urls = [
    "https://rn.veevavault.help/en/gr/archive/24r1/24r1-data-model-changes-quality/",
    "https://example.com/other-url"
]

for url in urls:
    data = fetch_and_scrape(url)
    if data:
        print(f"Data from {url}:")
        for item in data:
            print(f"Heading: {item['heading']}")
            print("Paragraphs:")
            for paragraph in item['content']['paragraphs']:
                print(f"  - {paragraph}")
            print("Lists:")
            for list_item in item['content']['lists']:
                print(f"  - {list_item}")
    else:
        print(f"No data fetched for {url}.")       

        # Generate recommendations function based on Name and Results columns
        def generate_recommendations_based_on_results(filtered_df, search_column='Name'):
            recommendations = []
            for idx, row in filtered_df.iterrows():
                name = row.get(search_column, '').lower()  # Get the product name
                results_text = row.get('results', '')  # Get the associated search results

                # Check if both Name and Results contain data
                if name and results_text:
                    # Combine Name and Results for recommendation generation
                    prompt = f"""
                    You are an AI analyst. Below is a product name and its associated search results:

                    Product Name: {name}

                    Search Results:
                    {results_text}

                    Based on the search results, summarize, Keep the recommendation concise in 2-3 lines. Additionally, can you recommend the changes that need to be incorporated for the client?

                    Return the result in the following format:
                    {{
                        "Recommendation": "Recommendation text here"
                    }}
                    """


                    try:
                        # Send the prompt to Gemini Pro (or your AI model) for processing
                        response = chat.send_message(prompt, stream=False)

                        # Convert response to a single string
                        response_text = "".join([chunk.text for chunk in response])

                        # Parse the response as JSON
                        result = json.loads(response_text)

                        # Check if the response contains the recommendation
                        if isinstance(result, dict) and "Recommendation" in result:
                            recommendations.append(result["Recommendation"])
                        else:
                            recommendations.append("Error in AI response.")

                    except json.JSONDecodeError as e:
                        recommendations.append(f"JSON parsing error: {str(e)}")
                    except Exception as e:
                        recommendations.append(f"Error in processing: {str(e)}")
                else:
                    recommendations.append("No valid Name or Results found.")

            return recommendations


        # AI Assessment Function for Feature Descriptions
        def analyze_dataset_with_ai(filtered_df):
            # Ensure 'Feature Description' column exists
            if 'Feature Description' not in filtered_df.columns:
                st.error("The dataset must contain a 'Feature Description' column for analysis.")
                return []

            results = []
            for index, row in filtered_df.iterrows():
                feature_description = row['Feature Description']

                prompt = f"""
                You are an AI analyst. Below is a product feature description:

                {feature_description}

                1. Provide a brief assessment comment on the latest updates in Veeva Vault, highlighting any potential impact on the system or business operations. Ensure the comment helps the client understand the significance of the update in 2-3 concise lines

                Return the result in the following format:
                {{
                    "Assessment_comment": "Assessment text here"
                }}
                """

                try:
                    # Send the prompt to Gemini Pro (or your AI model) for processing
                    response = chat.send_message(prompt, stream=False)

                    # Convert response to a single string
                    response_text = "".join([chunk.text for chunk in response])

                    # Debugging: Log the raw AI response
                    st.write(f"Raw AI Response for {feature_description}: {response_text}")

                    # Parse the response as JSON
                    result = json.loads(response_text)

                    if isinstance(result, dict) and "Assessment_comment" in result:
                        results.append({
                            "row_index": index,
                            "assessment_comment": result["Assessment_comment"]
                        })
                    else:
                        st.error(f"Invalid response format for row {index}. Expected a dictionary with 'Assessment_comment'.")
                        results.append({
                            "row_index": index,
                            "assessment_comment": "Error in AI response"
                        })

                except json.JSONDecodeError as e:
                    st.error(f"JSON parsing error: {str(e)}")
                    results.append({
                        "row_index": index,
                        "assessment_comment": "Error in parsing AI response"
                    })
                except Exception as e:
                    st.error(f"Error in processing row {index}: {str(e)}")
                    results.append({
                        "row_index": index,
                        "assessment_comment": "Error in processing"
                    })

            return results

        # Streamlit App UI
        st.title("AI-Driven Feature Assessment and Web Scraping Recommendations")

        # Web scraping URL input
        # Filter out any NaN values from the 'extracted_link' column and get the first valid URL
        url_input = filtered_df['extracted_link'].dropna().iloc[0] if not filtered_df['extracted_link'].dropna().empty else ""

        # Display the link in the text input field
        url_input = st.text_input(
            "Enter the URL to scrape:", 
            url_input
        )


        # Combined Button to Trigger Both AI Assessment and Recommendations
        if st.button("Get Feature Assessments and Recommendations"):
            if filtered_df is not None and url_input:
                # Trigger AI analysis for feature descriptions
                ai_results = analyze_dataset_with_ai(filtered_df)

                # Web Scraping and Recommendation Generation
                scraped_data = scrape_data(url_input)

                if scraped_data:
                    st.subheader("Search Results:")
                    for idx, row in filtered_df.iterrows():
                        matched_results = []
                        name = row.get('Name', '').lower()

                        # Match the scraped content with the 'Name' column
                        for item in scraped_data:
                            heading = item['heading']
                            paragraphs = item['content']['paragraphs']
                            lists = item['content']['lists']

                            if name in heading.lower() or any(name in para.lower() for para in paragraphs) or any(name in lst.lower() for lst in lists):
                                matched_results.append(f"Found in: {heading}")
                                matched_results.extend(paragraphs)
                                matched_results.extend(lists)

                        filtered_df.at[idx, 'results'] = "\n".join(matched_results)

                    # Generate recommendations based on results
                    recommendations = generate_recommendations_based_on_results(filtered_df)
                    filtered_df['Recommendation'] = recommendations

                    # Add AI assessment results
                    for ai_result in ai_results:
                        filtered_df.at[ai_result["row_index"], 'Assessment_comment'] = ai_result["assessment_comment"]

                    # Display the updated DataFrame with both recommendations and assessments
                    st.write("Dataset with Assessments and Recommendations:", filtered_df)
                    
                    
                    #final dataset columns
                    
                    columns=['Application Family','Application','Name','Enablement Setting','GxP Risk','Default Impact',
                             'Feature Description','Release Note','extracted_link','Assessment_comment','Recommendation']
                    #final dataset 
                    st.write("Final dataset",filtered_df[columns])

                    # Option to download the updated DataFrame
                    csv = filtered_df.to_csv(index=False).encode('utf-8')
                    st.download_button("Download Updated CSV", csv, "updated_data.csv", "text/csv")
                else:
                    st.error("Failed to retrieve data from URL. Please check the URL or try again later.")
            else:
                st.warning("Please provide a URL to scrape.")
        else:
            st.warning("Please provide a URL to scrape.")

else:
    st.warning("Please upload a data file to get started.")
